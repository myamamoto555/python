# -*- coding: utf-8 -*-
import openpyxl as px

wb = px.load_workbook(filename='../datas/zap.xlsx')
sheetnames = wb.get_sheet_names()  # シートネーム一覧を取得

foutin = open("../datas/input.txt", "w")
foutout = open("../datas/output.txt", "w")
for sn in sheetnames:
    ws = wb.get_sheet_by_name(sn)  # 各シートネーム中の情報を取得
    firstflag = 0
    for i in range(3, 50000):
        speaker = ws['F'+str(i)].value
        session_id = ws['B'+str(i)].value
        if firstflag == 0:
            if speaker == 2:
                question = str(ws['G'+str(i)].value)
                firstflag = 1
                before_speaker = 2
                before_session_id = ws['B'+str(i)].value
        else:
            if before_session_id == session_id:
                if before_speaker == 2 and speaker == 1:
                    answer = str(ws['G'+str(i)].value)
                    before_speaker = 1
                elif before_speaker == 2 and speaker == 2:
                    question += "nl" + str(ws['G'+str(i)].value)
                    before_speaker = 2
                elif before_speaker == 1 and speaker == 2:
                    foutin.write(str(question).replace(" ", "") + "\n")
                    foutout.write(str(answer).replace(" ", "") + "\n")
                    question = str(ws['G'+str(i)].value)
                    before_speaker = 2
                elif before_speaker == 1 and speaker == 1:
                    answer += "nl" + str(ws['G'+str(i)].value)
                    before_speaker = 1
            else:
                foutin.write(str(question).replace(" ", "") + "\n")
                foutout.write(str(answer).replace(" ", "") + "\n")
                if speaker == 1:
                    firstflag = 0
                if speaker == 2:
                    question = str(ws['G'+str(i)].value)
                    before_speaker = 2
            before_session_id = ws['B'+str(i)].value

foutin.close()
foutout.close()
